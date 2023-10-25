import json
from openpyxl import load_workbook

# Carregar o arquivo Excel
xlsx_file = "relacao-contas.xlsx"
workbook = load_workbook(xlsx_file)
worksheet = workbook.active

# Criar um dicionário para armazenar os dados no formato desejado
data = {}

# Iterar pelas linhas da planilha (começando da segunda linha, assumindo que a primeira é o cabeçalho)
for row in worksheet.iter_rows(min_row=2, values_only=True):
    conta = {
        "email": row[0],
        "senha": row[1]
    }
    data[f"conta{len(data) + 1}"] = conta

# Escrever os dados no arquivo JSON
with open("relacao-contas.json", "w") as json_file:
    json.dump(data, json_file, indent=4)

print("Arquivo JSON 'relacao-contas.json' criado com sucesso!")